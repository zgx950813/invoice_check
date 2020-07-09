#!/usr/bin/python
# -*- coding: UTF-8 -*-

import sys

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from invoice import Invoice, Result
from web_writer import init_driver, input_invoice_info, handle_response


def check_invoice(chromedriver_path, wb_name, new_wb_name, begin_row=None, end_row=None):
    wb = load_workbook(filename=wb_name)
    ws = wb['发票查询']
    print('[INFO] 工作簿最大行数：%s' % ws.max_row)

    # 初始化环境
    driver = init_driver(chromedriver_path)

    # 判断输入合法性
    if begin_row != None and end_row != None and begin_row > end_row:
        begin_row, end_row = end_row, begin_row
    begin_row = int(2) if begin_row == None or int(
        begin_row) < 2 else int(begin_row)
    end_row = int(ws.max_row) if end_row == None or int(
        end_row) > int(ws.max_row) else int(end_row)
    print('[INFO] excel将从第 %s 行执行至第 %s 行' % (begin_row, end_row))

    # 开始执行
    for row in range(begin_row, end_row + 1):
        print('[INFO] 当前位于：第 %s 行' % row)
        # 读取excel中的数据
        vin = str(ws.cell(row=row, column=column_index_from_string(
            'B')).value)                         # 车架号码
        invoice_code = str(ws.cell(row=row, column=column_index_from_string(
            'C')).value).rjust(12, '0')  # 发票代码，总共12位，不足左边补0
        invoice_num = str(ws.cell(row=row, column=column_index_from_string(
            'D')).value)                 # 发票号码
        date_issued = str(ws.cell(row=row, column=column_index_from_string(
            'E')).value)                 # 开票日期
        check_price = str(ws.cell(row=row, column=column_index_from_string(
            'F')).value)                 # 不含税价（原车销售发票）/车价合计（二手车销售发票）
        invoice = Invoice(vin, invoice_code, invoice_num,
                          date_issued, check_price)
        invoice.print_info()

        # 输入发票信息
        input_invoice_info(driver, invoice)
        # 抓取键盘输入
        flag = True
        while True:
            key_input = input()
            if key_input == 'ok' or key_input == 'OK':
                print('[INFO] 程序继续执行')
                break
            elif key_input == 'exit' or key_input == 'EXIT':
                print('[INFO] 手动停止执行')
                flag = False
                break

        if flag == False:
            break

        # 处理网页返回的信息
        result = handle_response(driver, invoice)
        result.print_info()

        # 存放数据入excel
        ws.cell(row=row, column=column_index_from_string('G'), value=str(
            result.check_time)[str(result.check_time).find('：') + 1:])          # 查询时间
        ws.cell(row=row, column=column_index_from_string('H'), value=str(
            result.invoice_status)[str(result.invoice_status).find('：') + 1:])  # 发票状态
        ws.cell(row=row, column=column_index_from_string('I'), value=str(
            result.invoice_code))                                                # 发票代码
        ws.cell(row=row, column=column_index_from_string('J'), value=str(
            result.invoice_num))                                                 # 发票号码
        ws.cell(row=row, column=column_index_from_string('K'), value=str(
            result.date_issued))                                                 # 开票日期
        ws.cell(row=row, column=column_index_from_string('L'), value=str(
            result.check_price)[str(result.check_price).find('￥') + 1:])        # 不含税价（原车销售发票）/车价合计（二手车销售发票）
        ws.cell(row=row, column=column_index_from_string('M'), value=str(
            result.vin))                                                         # 车架号码
        ws.cell(row=row, column=column_index_from_string('N'), value=str(
            result.buyer_entity))                                                # 买方单位
        ws.cell(row=row, column=column_index_from_string('O'), value=str(
            result.buyer_identity))                                              # 买方证号
        ws.cell(row=row, column=column_index_from_string('P'), value=str(
            result.seller_entity))                                               # 卖方单位
        ws.cell(row=row, column=column_index_from_string('Q'), value=str(
            result.seller_identity))                                             # 卖方证号
        print('[INFO] 已完成：%.2f%%' %(100 * (row - begin_row + 1) / (end_row - begin_row + 1)))

    # 保存excel
    print('[INFO] excel已保存')
    wb.save(new_wb_name)
    # 退出浏览器
    driver.quit()
    print('[INFO] 欢迎再次使用！Kimmy (✿◡‿◡)')


if __name__ == "__main__":
    chromedriver_path = 'chromedriver.exe'
    wb_name = '半自动化发票查询实现目标.xlsx'
    new_wb_name = '半自动化发票查询结果.xlsx'

    if len(sys.argv) == 1:
        check_invoice(chromedriver_path, wb_name, new_wb_name)
    elif len(sys.argv) == 2:
        check_invoice(chromedriver_path, wb_name, new_wb_name, sys.argv[1])
    elif len(sys.argv) == 3:
        check_invoice(chromedriver_path, wb_name,
                      new_wb_name, sys.argv[1], sys.argv[2])
    else:
        exit(-1)
